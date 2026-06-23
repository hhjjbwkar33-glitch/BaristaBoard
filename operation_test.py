import pandas as pd
from datetime import datetime, timedelta

df = pd.read_excel('baristaboard_input.xlsx')

def split_shift(start, end):
    start = str(start)
    end = str(end)
    split_time = []
    start_time = datetime.strptime(start, '%H:%M:%S')
    end_time = datetime.strptime(end, '%H:%M:%S')
    current_time = start_time
    while current_time < end_time:
        next_time = current_time + timedelta(minutes=30)
        time_interval = current_time.strftime('%H:%M') + '-' + next_time.strftime('%H:%M')
        split_time.append(time_interval)
        current_time = next_time
    return split_time

def generate30x2break(schedule, start, end):
    start_time = datetime.strptime(start, '%H:%M:%S')
    end_time = datetime.strptime(end, '%H:%M:%S')
    break_30slot = []
    for slot, names in schedule.items():
        if len(names) >= 3:
            fs_start, fs_end = slot.split('-')
            fs_start = datetime.strptime(fs_start, '%H:%M')
            fs_end = datetime.strptime(fs_end, '%H:%M')
            if fs_start >= start_time + timedelta(hours=2) and fs_start < end_time - timedelta(hours=2):
                break_30slot.append((fs_start, fs_end))
    if not break_30slot:
        return None
    half = len(break_30slot) // 2
    quarter = half // 2
    start_break1, end_break1 = break_30slot[quarter]
    start_break2, end_break2 = break_30slot[half + quarter]
    return [(start_break1, end_break1, 30), (start_break2, end_break2, 30)]

def generate_break(start, end, schedule):
    start = str(start)
    end = str(end)
    work_time = datetime.strptime(end, '%H:%M:%S') - datetime.strptime(start, '%H:%M:%S')
    if work_time < timedelta(hours=5):
        return None
    elif work_time < timedelta(hours=6):
        break_time = 15
    elif work_time < timedelta(hours=7):
        break_time = 30
    elif work_time < timedelta(hours=8):
        start_break = datetime.strptime(start, '%H:%M:%S') + timedelta(hours=3)
        return [
            (start_break, start_break + timedelta(minutes=30), 30),
            (start_break + timedelta(minutes=30), start_break + timedelta(minutes=45), 15)
        ]
    else:
        result = generate30x2break(schedule, start, end)
        if result is None:
            print('休憩不可')
            return None
        return result
    start_break = datetime.strptime(start, '%H:%M:%S') + timedelta(hours=3)
    end_break = start_break + timedelta(minutes=break_time)
    return (start_break, end_break, break_time)

def is_peak(slot):
    peak_slots = {'12:00-12:30', '12:30-13:00', '13:00-13:30', '13:30-14:00'}
    return slot in peak_slots

def required_positions(num_people):
    if num_people == 2:
        return ['バリスタ', 'キャッシャー']
    else:
        return ['バリスタ', 'キャッシャー', 'フロア']

def can_assign(name, position, current_limit, used_count, last_position, streak_count, role_limit):
    if used_count.get(position, 0) >= current_limit[position]:
        return False
    if last_position.get(name) == position:
        if streak_count.get(name, 0) >= role_limit:
            return False
    return True

def check_shortage(slot, assigned_list, num_people, peak):
    if peak:
        required = {'バリスタ': 2, 'キャッシャー': 1, 'フロア': 1}
    elif num_people == 2:
        required = {'バリスタ': 1, 'キャッシャー': 1}
    else:
        required = {'バリスタ': 1, 'キャッシャー': 1}
    actual = {}
    for name, position in assigned_list:
        actual[position] = actual.get(position, 0) + 1
    shortages = []
    for position, required_count in required.items():
        actual_count = actual.get(position, 0)
        if actual_count < required_count:
            shortages.append({
                'slot': slot,
                'position': position,
                'required': required_count,
                'actual': actual_count,
                'shortage': required_count - actual_count
            })
            print(f"⚠ 欠員発生: {slot} | ポジション: {position} | 不足: {required_count - actual_count}")
    return shortages

# 最初にファイルを新規作成しておく
with pd.ExcelWriter('baristaboard_output.xlsx', engine='openpyxl') as writer:
    pd.DataFrame().to_excel(writer, sheet_name='Sheet1')
from openpyxl import Workbook
wb = Workbook()
wb.save('baristaboard_output.xlsx')
for date, group in df.groupby('日付'):
    schedule = {}
    break_schedule = {}
    records = []
    shortage_log = []
    role_limit = 4
    last_position = {}
    streak_count = {}
    new_schedule = {}
    position_limit = {
        '通常': {'バリスタ': 1, 'キャッシャー': 1, 'フロア': 1},
        'ピーク': {'バリスタ': 2, 'キャッシャー': 1, 'フロア': 1}
    }

    for index, row in group.iterrows():
        name = row['名前']
        start = row['開始時間']
        end = row['終了時間']
        slots = split_shift(start, end)
        for slot in slots:
            if slot in schedule:
                schedule[slot].append(name)
            else:
                schedule[slot] = [name]

    for index, row in group.iterrows():
        name = row['名前']
        start = row['開始時間']
        end = row['終了時間']
        result = generate_break(start, end, schedule)
        if result is not None:
            if name not in break_schedule:
                break_schedule[name] = []
            if isinstance(result, list):
                break_schedule[name].extend(result)
            else:
                break_schedule[name].append(result)

    for slot, names in schedule.items():
        assigned = []
        assigned_people = set()
        num_people = len(names)
        required = required_positions(num_people)
        current_limit = position_limit['ピーク'] if is_peak(slot) else position_limit['通常']
        used_count = {p: 0 for p in current_limit}

        for position in required:
            for name in names:
                if name in assigned_people:
                    continue
                slot_start, slot_end = slot.split('-')
                slot_start = datetime.strptime(slot_start, '%H:%M')
                slot_end = datetime.strptime(slot_end, '%H:%M')
                is_on_break = False
                for break_start, break_end, _ in break_schedule.get(name, []):
                    if break_start < slot_end and break_end > slot_start:
                        is_on_break = True
                        break
                if is_on_break:
                    continue
                if name not in last_position:
                    last_position[name] = position
                    streak_count[name] = 0
                if can_assign(name, position, current_limit, used_count, last_position, streak_count, role_limit):
                    if last_position[name] == position:
                        streak_count[name] += 1
                    else:
                        streak_count[name] = 1
                    last_position[name] = position
                    used_count[position] += 1
                    assigned.append((name, position))
                    assigned_people.add(name)
                    break

        new_schedule[slot] = assigned

    for slot, assigned in new_schedule.items():
        for name, position in assigned:
            records.append({'時間': slot, '名前': name, 'ポジション': position})
        num_people = len(schedule.get(slot, []))
        peak = is_peak(slot)
        shortages = check_shortage(slot, assigned, num_people, peak)
        shortage_log.extend(shortages)

    result_df = pd.DataFrame(records)
    pivot_df = result_df.pivot(index='時間', columns='名前', values='ポジション')
    pivot_df = pivot_df.T
    pivot_df = pivot_df.replace({'バリスタ': 'B', 'キャッシャー': 'C', 'フロア': 'F'})
    pivot_df = pivot_df.fillna('ー')

    for name, breaks in break_schedule.items():
        for break_start, break_end, break_time in breaks:
            for slot in pivot_df.columns:
                slot_start, slot_end = slot.split('-')
                slot_start = datetime.strptime(slot_start, '%H:%M')
                slot_end = datetime.strptime(slot_end, '%H:%M')
                break_minutes = int((break_end - break_start).total_seconds() / 60)
                if slot_end > break_start and slot_start < break_end:
                    pivot_df.loc[name, slot] = f'休憩({break_minutes})'

    sheet_name = pd.Timestamp(date).strftime('%m月%d日')

    with pd.ExcelWriter('baristaboard_output.xlsx', mode='a', if_sheet_exists='overlay') as writer:
        pivot_df.to_excel(writer, sheet_name=sheet_name, index=True)
        if shortage_log:
            shortage_df = pd.DataFrame(shortage_log)
            # シフト表の下に欠員ログを追記
            startrow = len(pivot_df) + 3
            print(f'startrow: {startrow}')
        
        shortage_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

    from openpyxl import load_workbook
    wb = load_workbook('baristaboard_output.xlsx')
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save('baristaboard_output.xlsx')

print('出力完了！')